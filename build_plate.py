

"""
Build Plates
"""

# standard libraries

# nonstandard libraries
import openpyxl
from Bio.Seq import Seq
from Bio.SeqUtils import MeltingTemp as mt

# homegrown libraries

def get_cas9_sites(gene_dicts,*args,**kwargs):
    # get default settings
    settings = {
            "min_spacing":10,
            "5' spacing":10,
            "3' spacing":0,
            "gRNA length":19,
            "PAM":"NGG",
            }
    # update settings dictionary
    for arg in args: settings.update(arg)
    settings.update(kwargs)
    # transfer to local namespace
    pam = settings["PAM"]
    gRNA_length = settings["gRNA length"]
    min_spacing = settings["min_spacing"]
    p5_spacing = settings["5' spacing"]
    p3_spacing = settings["3' spacing"]
    # 
    for gene,gene_dict in gene_dicts.items():
        print 'Starting analysis on {}...'.format(gene)

        for exon_set in gene_dict:

            exon,upstream,downstream = [exon_set[l] for l in ('exon','upstream','downstream')]

            print '\nNew exon:'

            segment = upstream + exon + downstream # piece together target
            ind = len(upstream + pam) + p5_spacing

            while ind < len(upstream + exon) - len(pam) - p3_spacing + 1:

                if check_pam(pam,segment[ind:ind + len(pam)]):
                    print segment[ind-gRNA_length:ind+len(pam)]

                ind += 1


        



def load_sequences(fname):
    
    # initialize variables
    sequences = {}

    # load worksheet
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    
    # extract rows from active sheet
    rows = list(iter_rows(ws))
    
    # get indices for each column
    header = [str(s).lower() for s in rows[0]]
    gene_ind = header.index('gene')
    exon_ind = header.index('exon')
    upstream_ind = header.index('upstream')
    downstream_ind = header.index('downstream')

    # pull out genes
    genes = list(set([str(row[gene_ind]) for row in rows[1:]]))

    # create list for each gene
    for gene in genes:
        sequences[gene] = []
    
    # iterate through rows and plop items in
    for row in rows[1:]:
        sequences[row[gene_ind]].append({
            'exon':      interpret(row[exon_ind]),
            'upstream':  interpret(row[upstream_ind]),
            'downstream':interpret(row[downstream_ind]),
            })

    return sequences

""" ---------------- """
""" HELPER FUNCTIONS """
""" ---------------- """

def get_dna_code():
    """ Return dictionary of degenerative codons """
    return dict([
        ('A','A'), ('T','T'), ('C','C'), ('G','G'),
        ('B','CGT'), ('D','AGT'), ('H','ACT'), ('K','GT'),
        ('M','AC'), ('N','ACGT'), ('R','AG'), ('S','CG'),
        ('V','ACG'), ('W','AT'), ('Y','CT')])

def interpret(my_str):
    """ Swaps None for empty str """
    if my_str is None:
        return ''
    return my_str
        

def check_pam(pam,seq,dna_code = {}):

    # load DNA translation code
    dna_code = get_dna_code()

    # check length of sequences is the same
    if len(pam) != len(seq):
        print pam,seq
        raise ValueError('Lengths not equal!')
    
    # iterate through letters
    for p,s in zip(pam.upper(),seq.upper()):
        if not s in dna_code[p]:
            return False

    # if no template breaking matches
    return True


""" Yields rows from sheet """
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]





if __name__ == "__main__":
    sequences = load_sequences('TCR-Constant.xlsx')
    get_cas9_sites(sequences)












